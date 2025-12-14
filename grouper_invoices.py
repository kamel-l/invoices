import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Any

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class StylesFacture:
    """Classe définissant les styles visuels pour les factures"""
    BLEU_CLAIR = "CCCCCC"
    #BLEU_CLAIR = "E3F2FD"
    BLEU_FONCE = "1976D2"
    GRIS_CLAIR = "F5F5F5"
    GRIS_FONCE = "616161"
    VERT_CLAIR = "E8F5E9"
    VERT_FONCE = "2E7D32"
    CLAIR_PISTACHIO = "CCCCCC"

    titre_principal = Font(name='Calibri', size=16, bold=True, color=BLEU_FONCE)
    sous_titre = Font(name='Calibri', size=12, bold=True, color=GRIS_FONCE)
    texte_normal = Font(name='Calibri', size=11)

    bordure_complete = Border(
        left=Side(style='thin', color=GRIS_FONCE),
        right=Side(style='thin', color=GRIS_FONCE),
        top=Side(style='thin', color=GRIS_FONCE),
        bottom=Side(style='thin', color=GRIS_FONCE)
    )

    bordure_epaisse = Border(
        left=Side(style='medium', color=BLEU_FONCE),
        right=Side(style='medium', color=BLEU_FONCE),
        top=Side(style='medium', color=BLEU_FONCE),
        bottom=Side(style='medium', color=BLEU_FONCE)
    )

    remplissage_entete = PatternFill(start_color=BLEU_CLAIR, end_color=BLEU_CLAIR, fill_type='solid')
    remplissage_total = PatternFill(start_color=VERT_CLAIR, end_color=VERT_CLAIR, fill_type='solid')

class FormateurFacture:
    """Classe pour appliquer le formatage aux factures"""
    def __init__(self):
        self.styles = StylesFacture()

    def appliquer_style_enter(self, worksheet, row: int, col_start: int, col_end: int):
        """Applique le style d'en-tête à une rangée"""
        for col in range(col_start, col_end + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = self.styles.sous_titre
            cell.fill = self.styles.remplissage_entete
            cell.border = self.styles.bordure_complete
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def appliquer_style_ligne(self, worksheet, row: int, col_start: int, col_end: int):
        """Applique le style de ligne standard"""
        for col in range(col_start, col_end + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = self.styles.texte_normal
            cell.border = self.styles.bordure_complete
            cell.alignment = Alignment(vertical='center')

    def formater_facture(self, worksheet):
        """Formate une feuille de facture complète"""
        worksheet['A1'] = 'FACTURE'
        worksheet['A1'].font = self.styles.titre_principal
        worksheet.merge_cells('A1:F1')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        self.appliquer_style_enter(worksheet, 2, 1, 6)
        for row in range(3, 6):
            self.appliquer_style_ligne(worksheet, row, 1, 6)

        self.appliquer_style_enter(worksheet, 7, 1, 6)
        dernier_ligne = worksheet.max_row

        for row in range(8, dernier_ligne - 4):
            self.appliquer_style_ligne(worksheet, row, 1, 6)

        for row in range(dernier_ligne - 3, dernier_ligne - 1):
            cell = worksheet.cell(row=row, column=1)
            cell.font = self.styles.sous_titre
            cell.fill = self.styles.remplissage_total
            cell.border = self.styles.bordure_epaisse

        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            worksheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4

class GestionnaireFactures:
    def __init__(self, dossier_source: str, fichier_destination: str):
        self.dossier_source = Path(dossier_source)
        self.fichier_destination = Path(fichier_destination)
        self.donnees: Dict[str, Any] = {}
        self.formateur = FormateurFacture()

    def charger_fichiers(self) -> None:
        """Charge tous les fichiers Excel du dossier source"""
        if not self.dossier_source.exists():
            raise FileNotFoundError(f"Le dossier {self.dossier_source} n'existe pas")

        fichiers_excel = list(self.dossier_source.glob("*.xlsx"))

        if not fichiers_excel:
            logger.warning(f"Aucun fichier Excel trouvé dans {self.dossier_source}")
            return

        for fichier in fichiers_excel:
            try:
                classeur = pd.read_excel(fichier, sheet_name=None)
                self.donnees.update(classeur)
                logger.info(f"Fichier chargé : {fichier}")
            except Exception as e:
                logger.error(f"Erreur lors de la lecture de {fichier}: {e}")

    def sauvegarder_fusion(self) -> None:
        """Sauvegarde et formate toutes les données dans un fichier unique"""
        if not self.donnees:
            logging.warning("Aucune donnée à sauvegarder")
            return

        try:
            with pd.ExcelWriter(self.fichier_destination, engine='openpyxl') as writer:
                for nom_feuille, donnees_feuille in self.donnees.items():
                    donnees_feuille.to_excel(writer, sheet_name=nom_feuille, index=False)
                    worksheet = writer.sheets[nom_feuille]
                    self.formateur.formater_facture(worksheet)

            logging.info(f"Fichiers fusionnés et formatés dans : {self.fichier_destination}")
        except Exception as e:
            logging.error(f"Erreur lors de la fusion des fichiers : {e}")
            raise

def main():
    """Fonction principale"""
    try:
        dossier_source = 'invoices24/SIDALI_DINOTEX'
        fichier_destination = 'invoices24/INVOICES_SIDALI_DINOTEX_24.xlsx'

        gestionnaire = GestionnaireFactures(dossier_source, fichier_destination)
        gestionnaire.charger_fichiers()
        gestionnaire.sauvegarder_fusion()

        print("Traitement terminé avec succès")
    except Exception as e:
        print(f"Erreur lors du traitement : {str(e)}")
        raise

if __name__ == "__main__":
    main()
