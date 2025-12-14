import os
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
from urllib.parse import unquote

# Définir le chemin vers le dossier contenant les classeurs à regrouper
dossier_source = r'invoices_excel0/0661651722'

# Définir le nom du classeur de destination
classeur_destination = 'classeur_final.xlsx'


def sauvegarder_excel(fichier_name='facture.xlsx'):
    Date = pairs.get("Date", "")
    # Créer un nouveau classeur Excel
    writer = pd.ExcelWriter(fichier_name, engine='openpyxl')

    # Créer le DataFrame pour les informations client
    df_client = pd.DataFrame([resume['informations_client']])
    df_client.to_excel(writer, sheet_name=Date, startrow=1, header=True, index=False)

    # Créer le DataFrame pour les articles
    df_articles = pd.DataFrame(resume['articles'])
    df_articles.to_excel(writer, sheet_name=Date, startrow=6, header=True, index=False)

    # Obtenir la feuille de calcul
    workbook = writer.book
    worksheet = writer.sheets[Date]

    # Styles
    header_style = Font(bold=True, size=12)
    cell_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    header_fill = PatternFill(start_color='DFFF00', end_color='DFFF00', fill_type='solid')

    # Formater l'en-tête
    worksheet['A1'] = 'FACTURE'
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet.merge_cells('A1:F1')
    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet['A1'].color = Color(rgb=(67, 255, 100))

    # Formater les informations client
    for col in range(1, df_client.shape[1] + 1):
        cell = worksheet.cell(row=2, column=col)
        cell.font = header_style
        cell.fill = header_fill
        cell.border = cell_border

    # Formater les articles
    for col in range(1, df_articles.shape[1] + 1):
        cell = worksheet.cell(row=7, column=col)
        cell.font = header_style
        cell.fill = header_fill
        cell.border = cell_border

    # Ajouter les totaux
    row_totaux = 8 + len(resume['articles'])
    worksheet.cell(row=row_totaux, column=1, value='Total HT:')
    worksheet.cell(row=row_totaux, column=2, value=resume['totaux']['total_ht'])
    worksheet.cell(row=row_totaux + 1, column=1, value='Montant payé:')
    worksheet.cell(row=row_totaux + 1, column=2, value=resume['totaux']['montant_paye'])

    # Ajouter les notes
    row_notes = row_totaux + 3
    worksheet.cell(row=row_notes, column=1, value='Notes:')
    worksheet.cell(row=row_notes, column=2, value=resume['notes']['notes_communes'])
    worksheet.cell(row=row_notes + 1, column=1, value='Remarques:')
    worksheet.cell(row=row_notes + 1, column=2, value=resume['notes']['notes_pied'])

    # Ajuster la largeur des colonnes
    for col in worksheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            worksheet.column_dimensions[column].width = max_length + 2

    # Sauvegarder le fichier
    writer._save()
# Initialiser un dictionnaire pour stocker les données
donnees = {}

# Parcourir les fichiers dans le dossier source
for fichier in os.listdir(dossier_source):
    if fichier.endswith('.xlsx'):
        # Construire le chemin complet du fichier
        chemin_fichier = os.path.join(dossier_source, fichier)

        # Lire les feuilles du classeur dans un dictionnaire
        classeur = pd.read_excel(chemin_fichier, sheet_name=None)
        print(classeur)

        # Ajouter les données du classeur au dictionnaire principal
        donnees.update(classeur)

# Créer un nouveau classeur de destination et écrire les données
with pd.ExcelWriter(classeur_destination) as writer:
    for nom_feuille, donnees_feuille in donnees.items():
        donnees_feuille.to_excel(writer, sheet_name=nom_feuille, index=False)

print(f"Les feuilles ont été regroupées dans le classeur '{classeur_destination}'.")