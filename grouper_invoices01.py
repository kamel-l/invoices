import logging
from pathlib import Path
from typing import Dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import re


# ========================================================
# LOGGING
# ========================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
log = logging.getLogger("FactureOrganizer")



# ========================================================
# STYLE EXCEL MODERNE
# ========================================================
class ElegantExcelStyle:
    """Thème moderne, élégant et professionnel pour Excel"""

    def __init__(self):
        self._define()

    def _define(self):
        # Fonts
        self.title_font = Font("Segoe UI", 18, bold=True, color="2B2B2B")
        self.header_font = Font("Segoe UI", 11, bold=True, color="FFFFFF")
        self.text_font = Font("Segoe UI", 10, color="333333")
        self.amount_font = Font("Segoe UI", 10, color="1A5276")
        self.total_font = Font("Segoe UI", 11, bold=True, color="145A32")

        # Fills
        self.header_fill = PatternFill("solid", fgColor="2E86C1")
        self.even_fill = PatternFill("solid", fgColor="F4F6F7")
        self.total_fill = PatternFill("solid", fgColor="D4EFDF")

        # Borders
        thick = Side(style="thick", color="2C3E90")
        self.border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

        # Alignment
        self.center = Alignment(horizontal="center", vertical="center")
        self.left = Alignment(horizontal="left", vertical="center")
        self.right = Alignment(horizontal="right", vertical="center")


        # COLORATION DU ROW 4
        self.detail_header_fill = PatternFill("solid", fgColor="D6EA55")  # Rose léger
        self.detail_header_font = Font("Segoe UI", 11, bold=True, color="1B4F72")
        
        # COLORATION DU ROW 9
        self.row9_fill = PatternFill("solid", fgColor="D6EA55")  # bleu clair
        self.row9_font = Font("Segoe UI", 10, bold=True, color="1B4F72")


    def apply(self, cell, style="text"):
        if style == "title":
            cell.font = self.title_font
            cell.alignment = self.center

        elif style == "header":
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center

        elif style == "amount":
            cell.font = self.amount_font
            cell.alignment = self.right
            cell.number_format = "#,##0"

        elif style == "total":
            cell.font = self.total_font
            cell.fill = self.total_fill
            cell.alignment = self.right
            cell.number_format = "#,##0"

        else:
            cell.font = self.text_font
            cell.alignment = self.left



# ========================================================
# ORGANISATEUR DE FACTURES
# ========================================================
class InvoiceOrganizer:

    def __init__(self, source_dir: str, output_dir: str):
        self.source = Path(source_dir)
        self.output = Path(output_dir)
        self.output.mkdir(exist_ok=True)

        self.style = ElegantExcelStyle()

    # ----------------------------------------------------
    def process(self):
        clients = [d for d in self.source.iterdir() if d.is_dir()]
        if not clients:
            log.warning("Aucun dossier client trouvé.")
            return

        for client in clients:
            log.info(f"Traitement — {client.name}")
            data = self._read_client_files(client)
            if data:
                self._create_excel(client.name, data)

    # ----------------------------------------------------
    def _read_client_files(self, folder: Path) -> Dict[str, dict]:
        results = {}
        

        for file in folder.glob("*.xlsx"):
            try:
                wb = load_workbook(file)
                ws = wb.active

                date = self._extract_date(file.stem, ws)
                results[date] = {"ws": ws, "file": file.name}

                log.info(f"  ✓ {file.name}")

            except Exception as e:
                log.error(f"  ✗ Erreur {file.name} — {e}")

        return results
        
    # ----------------------------------------------------
    def _extract_date(self, filename: str, ws):
        patterns = [
            r"\d{4}[-/]\d{2}[-/]\d{2}",  # AAAA-MM-JJ
            r"\d{2}[-/]\d{2}[-/]\d{4}"   # JJ-MM-AAAA
        ]

        # 1) Chercher dans le nom du fichier
        for p in patterns:
            m = re.search(p, filename)
            if m:
                return m.group(0).replace("/", "-")

        # 2) Chercher dans les cellules
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    for p in patterns:
                        m = re.search(p, cell)
                        if m:
                            return m.group(0).replace("/", "-")

        # 3) Fallback
        return f"Inconnue_{datetime.now().strftime('%H%M%S')}"

    # ----------------------------------------------------
    def _unique_sheet_name(self, wb, base):
        base = base[:28]
        name = base
        i = 1
        while name in wb.sheetnames:
            name = f"{base}_{i}"
            i += 1
        return name

    # ----------------------------------------------------
    def _extract_totals(self, ws):
        keywords_ht = ["total ht", "hors taxes", "montant ht"]
        keywords_paid = ["payé", "paye", "reçu", "prélevé"]

        ht = paid = 0

        for row in ws.iter_rows(values_only=True):
            if not isinstance(row[0], str):
                continue

            label = row[0].lower()

            for k in keywords_ht:
                if k in label:
                    try:
                        ht = float(row[1])
                    except:
                        ht = 0

            for k in keywords_paid:
                if k in label:
                    try:
                        paid = float(row[1])
                    except:
                        paid = 0

        return {"ht": ht, "paid": paid, "balance": ht - paid}

    # ----------------------------------------------------
    def _create_excel(self, client, data: Dict[str, dict]):
        wb = Workbook()

        self._create_summary_sheet(wb, client, data)
        self._create_detail_sheets(wb, client, data)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        out = self.output / f"{client}.xlsx"
        wb.save(out)

        log.info(f"✔ Fichier créé : {out}")

    # ----------------------------------------------------
    def _create_summary_sheet(self, wb, client, data):
        ws = wb.create_sheet("Résumé")
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Factures – {client}"
        self.style.apply(ws["A1"], "title")

        ws.append(["Date", "Fichier", "Total HT", "Payé", "Solde", "Statut"])
        for col in range(1, 7):
            self.style.apply(ws.cell(row=2, column=col), "header")

        row = 3
        for date, info in data.items():
            totals = self._extract_totals(info["ws"])

            ws.append([
                date,
                info["file"],
                totals["ht"],
                totals["paid"],
                totals["balance"],
                "Payée" if totals["balance"] == 0 else "En attente"
            ])

            for col in range(1, 7):
                style = "amount" if col in (3, 4, 5) else "text"
                self.style.apply(ws.cell(row=row, column=col), style)

            row += 1

        self._auto_width(ws)

    # ----------------------------------------------------
    def _create_detail_sheets(self, wb, client,  data):
        for date, info in data.items():
            
            sheet_name = self._unique_sheet_name(wb, date)
            ws = wb.create_sheet(sheet_name)

            ws.merge_cells("A1:E1")
            ws["A1"] = f"Facture — {client}"
            self.style.apply(ws["A1"], "title")

            source_ws = info["ws"]
            start_row = 4

            # Copier les données
            for r_idx, row in enumerate(source_ws.iter_rows(values_only=True), start=start_row - 1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    kind = "amount" if isinstance(value, (int, float)) else "text"
                    self.style.apply(cell, kind)

            last_row = r_idx
            last_col = source_ws.max_column

            # ----------------------------------------------------
            # COLORATION DU ROW DES ENTÊTES (DESCRIPTION, CODE…)
            # ----------------------------------------------------

            header_row = start_row   # généralement ligne 4
            for col in range(1, last_col + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.fill = self.style.detail_header_fill
                cell.font = self.style.detail_header_font
                cell.alignment = self.style.center

            # ----------------------------------------------------
            # COLORATION DU ROW 9
            # ----------------------------------------------------

            row9 = 9
            for col in range(1, last_col + 1):
                cell = ws.cell(row=row9, column=col)
                # إذا تحب نفس اللون:
                # cell.fill = self.style.detail_header_fill
                # cell.font = self.style.detail_header_font

                # إذا تحب لون مختلف للصف 9:
                cell.fill = self.style.row9_fill
                cell.font = self.style.row9_font
                cell.alignment = self.style.center 

            # ====================================================
            # TABLEAU STRUCTURÉ
            # ====================================================
            table_ref = f"A{start_row}:{get_column_letter(last_col)}{last_row}"

            table = Table(
                displayName=f"Facture_{date.replace('-', '_')}",
                ref=table_ref
            )

            table_style = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False,
            )
            

            # Bordures épaisses
            for row in ws.iter_rows(min_row=start_row, max_row=last_row,
                                    min_col=1, max_col=last_col):
                for cell in row:
                    cell.border = self.style.border_thick
       
            
            
            # ====================================================
            # TOTAL AUTOMATIQUE – uniquement la colonne "Total"
            # ====================================================

            # نفترض أن عمود TOTAL هو آخر عمود في الجدول
            total_col = last_col  
            col_letter = get_column_letter(total_col)

            # موقع سطر التوتال
            total_row = last_row + 1

            # كتابة LABEL
            ws.cell(row=total_row, column=1, value="TOTAL")
            self.style.apply(ws.cell(row=total_row, column=1), "total")

            # فورمولا SUM للعمود TOTAL فقط
            formula = f"=SUM({col_letter}{start_row}:{col_letter}{last_row})"
            total_cell = ws.cell(row=total_row, column=total_col, value=formula)
            self.style.apply(total_cell, "total")

            # تلوين كامل السطر
            for col in range(1, last_col + 1):
                c = ws.cell(row=total_row, column=col)
                c.fill = self.style.total_fill
                c.font = self.style.total_font
                c.alignment = self.style.right if col == total_col else self.style.left

            
            table = Table(
                        displayName=f"Facture_{date.replace('-', '_')}",
                        ref=table_ref
                    )

            # 🔥 Enlever les en-têtes du tableau
            table.headerRowCount = 0

            table_style = TableStyleInfo(
                        name="TableStyleMedium9",
                        showRowStripes=True,
                        showColumnStripes=False,
                    )
            table.tableStyleInfo = table_style

            ws.add_table(table)


            # ====================================================
            # GRAPHIQUE DES PAIEMENTS
            # ====================================================
            

            self._auto_width(ws)

    # ----------------------------------------------------
    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            width = min(max_len + 2, 45)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width



# ========================================================
# MAIN
# ========================================================
def main():
    organizer = InvoiceOrganizer(
        source_dir="invoices25_downloaded",
        output_dir="factures25_modernes"
    )
    organizer.process()


if __name__ == "__main__":
    main()
