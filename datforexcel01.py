from urllib.parse import unquote
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datfilereader import DatFileReader
from pathlib import Path
import os


class FactureParser:
    def __init__(self, donnees_brutes):
        # Découper la chaîne en paires clé-valeur
        self.pairs = dict(pair.split('=') for pair in donnees_brutes.split('&'))
        # Décoder les valeurs URL-encodées
        self.pairs = {k: unquote(v) for k, v in self.pairs.items()}

    def obtenir_infos_client(self):
        return {
            "client": self.pairs.get("Customer", ""),
            "date": datetime.strptime(self.pairs.get("Date", ""), "%Y-%m-%d").strftime("%d/%m/%Y"),
            "delai_paiement": self.pairs.get("PaymentTermsDays", ""),
            "adresse": self.pairs.get("Address", ""),
            "ref_commande": self.pairs.get("CustomerPO", "")
        }

    def obtenir_articles(self):
        articles = []
        nb_articles = int(self.pairs.get("ItemCount", 0))

        for i in range(1, nb_articles + 1):
            article = {
                "description": self.pairs.get(f"Item{i}Description", ""),
                "code": self.pairs.get(f"Item{i}Code", ""),
                "quantite": float(self.pairs.get(f"Item{i}Qty", 0)),
                "prix_unitaire": float(self.pairs.get(f"Item{i}UnitValue", 0)),
                "remise": float(self.pairs.get(f"Item{i}Discount", 0)),
                "tva": self.pairs.get(f"Item{i}VATPercentage", "0%").replace("%", "")
            }

            # Calculer le total pour cet article
            article["total"] = article["quantite"] * article["prix_unitaire"]
            articles.append(article)

        return articles

    def obtenir_totaux(self):
        return {
            "total_ht": float(self.pairs.get("Total", 0)),
            "montant_paye": float(self.pairs.get("AmountPaid", 0)),
            "frais_expedition": float(self.pairs.get("ShippingCosts", 0))
        }

    def obtenir_notes(self):
        return {
            "notes": self.pairs.get("Notes", ""),
            "notes_communes": self.pairs.get("CommonNotes", ""),
            "notes_pied": self.pairs.get("CommonFootNotes", "")
        }


def sauvegarder_excel(fichier_name, resume, pairs):
    # Créer un nouveau classeur Excel
    writer = pd.ExcelWriter(fichier_name, engine='openpyxl')
    
    Date = pairs.get("Date", "")
    
    # Créer le DataFrame pour les informations client
    df_client = pd.DataFrame([resume['informations_client']])
    df_client.to_excel(writer, sheet_name=Date, startrow=1, header=True, index=False)

    # Créer le DataFrame pour les articles
    df_articles = pd.DataFrame(resume['articles'])
    df_articles.to_excel(writer, sheet_name=Date, startrow=6, header=True, index=False)

    # Obtenir la feuille de calcul
    workbook = writer.book
    worksheet = writer.sheets[Date]

    # Styles
    header_style = Font(bold=True, size=12)
    cell_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    header_fill = PatternFill(start_color='DFFF00', end_color='DFFF00', fill_type='solid')

    # Formater l'en-tête
    worksheet['A1'] = 'FACTURE'
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet.merge_cells('A1:F1')
    worksheet['A1'].alignment = Alignment(horizontal='center')

    # Formater les informations client
    for col in range(1, df_client.shape[1] + 1):
        cell = worksheet.cell(row=2, column=col)
        cell.font = header_style
        cell.fill = header_fill
        cell.border = cell_border

    # Formater les articles
    for col in range(1, df_articles.shape[1] + 1):
        cell = worksheet.cell(row=7, column=col)
        cell.font = header_style
        cell.fill = header_fill
        cell.border = cell_border

    # Ajouter les totaux
    row_totaux = 8 + len(resume['articles'])
    worksheet.cell(row=row_totaux, column=1, value='Total HT:')
    worksheet.cell(row=row_totaux, column=2, value=resume['totaux']['total_ht'])
    worksheet.cell(row=row_totaux + 1, column=1, value='Montant payé:')
    worksheet.cell(row=row_totaux + 1, column=2, value=resume['totaux']['montant_paye'])

    # Ajouter les notes
    row_notes = row_totaux + 3
    worksheet.cell(row=row_notes, column=1, value='Notes:')
    worksheet.cell(row=row_notes, column=2, value=resume['notes']['notes_communes'])
    worksheet.cell(row=row_notes + 1, column=1, value='Remarques:')
    worksheet.cell(row=row_notes + 1, column=2, value=resume['notes']['notes_pied'])

    # Ajuster la largeur des colonnes
    for col in worksheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        worksheet.column_dimensions[column].width = max_length + 2

    # Sauvegarder le fichier
    writer._save()


# Liste des fichiers trouvés
list_invoices = [nom_fichier for nom_fichier in os.listdir("invoices") if
                 os.path.isfile(os.path.join("invoices", nom_fichier))]

lecteur = DatFileReader()
for invoice in list_invoices:
    contenu = lecteur.lire_texte_simple(f'invoices/{invoice}')

    # Traiter la facture
    parser = FactureParser(contenu[0])
    resume = {
        'informations_client': parser.obtenir_infos_client(),
        'articles': parser.obtenir_articles(),
        'totaux': parser.obtenir_totaux(),
        'notes': parser.obtenir_notes()
    }
    
    pairs = dict(pair.split('=') for pair in contenu[0].split('&'))
    date = pairs.get("Date", "")
    client = pairs.get("Customer", "")
    
    # Nettoyer le nom du client pour qu'il soit valide comme nom de dossier
    client_clean = client.replace("%20", "_").replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace('"', "_").replace("<", "_").replace(">", "_").replace("|", "_")
    
    # Créer le dossier principal pour tous les clients
    download_path = Path("invoices_downloaded")
    download_path.mkdir(parents=True, exist_ok=True)
    
    # Créer le dossier spécifique pour ce client
    client_folder = download_path / client_clean
    client_folder.mkdir(parents=True, exist_ok=True)
    
    Ninvoice = invoice.replace(".dat", "")
    
    # Construire le chemin du fichier Excel dans le dossier du client
    nom_fichier = f"{client_clean}-{date}.xlsx"
    chemin_complet = client_folder / nom_fichier
    
    # Sauvegarder en Excel
    sauvegarder_excel(str(chemin_complet), resume, pairs)
    print(f"Facture sauvegardée avec succès pour {client} : {chemin_complet}")

print(f"\nTraitement terminé ! Les factures ont été organisées par client dans le dossier 'invoices_downloaded'")