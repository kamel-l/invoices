from urllib.parse import unquote
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datfilereader import DatFileReader



class FactureParser:
    def __init__(self, donnees_brutes):
        # Découper la chaîne en paires clé-valeur
        self.pairs = dict(pair.split('=') for pair in donnees_brutes.split('&'))
        # Décoder les valeurs URL-encodées
        self.pairs = {k: unquote(v) for k, v in self.pairs.items()}

    def obtenir_infos_client(self):
        return {
            "client": self.pairs.get("Customer", ""),
            "date": datetime.strptime(self.pairs.get("Date", ""), "%Y-%m-%d").strftime("%d/%m/%Y"),
            "delai_paiement": self.pairs.get("PaymentTermsDays", ""),
            "adresse": self.pairs.get("Address", ""),
            "ref_commande": self.pairs.get("CustomerPO", "")
        }

    def obtenir_articles(self):
        articles = []
        nb_articles = int(self.pairs.get("ItemCount", 0))

        for i in range(1, nb_articles + 1):
            article = {
                "description": self.pairs.get(f"Item{i}Description", ""),
                "code": self.pairs.get(f"Item{i}Code", ""),
                "quantite": float(self.pairs.get(f"Item{i}Qty", 0)),
                "prix_unitaire": float(self.pairs.get(f"Item{i}UnitValue", 0)),
                "remise": float(self.pairs.get(f"Item{i}Discount", 0)),
                "tva": self.pairs.get(f"Item{i}VATPercentage", "0%").replace("%", "")
            }

            # Calculer le total pour cet article
            article["total"] = article["quantite"] * article["prix_unitaire"]
            articles.append(article)

        return articles

    def obtenir_totaux(self):
        return {
            "total_ht": float(self.pairs.get("Total", 0)),
            "montant_paye": float(self.pairs.get("AmountPaid", 0)),
            "frais_expedition": float(self.pairs.get("ShippingCosts", 0))
        }

    def obtenir_notes(self):
        return {
            "notes": self.pairs.get("Notes", ""),
            "notes_communes": self.pairs.get("CommonNotes", ""),
            "notes_pied": self.pairs.get("CommonFootNotes", "")
        }


def sauvegarder_excel(resume, nom_fichier='facture.xlsx'):
    global date
    liste_invoices = ['10000.dat', '10001.dat', '10002.dat', '10003.dat', '10004.dat', '10005.dat', '10006.dat',
                      '10007.dat', '10008.dat', '10009.dat', '10010.dat', '10011.dat', '10012.dat', '10013.dat',
                      '10014.dat', '10015.dat', '10016.dat', '10017.dat', '10018.dat', '10019.dat', '10020.dat',
                      '10021.dat', '10022.dat', '10023.dat', '10024.dat', '10025.dat', '10027.dat', '10028.dat',
                      '10030.dat', '10032.dat', '10036.dat', '10037.dat', '10038.dat', '10039.dat', '10040.dat',
                      '10041.dat', '10042.dat', '10043.dat', '10044.dat', '10045.dat', '10046.dat', '10047.dat',
                      '10048.dat', '10049.dat', '10050.dat', '10051.dat', '10052.dat', '10053.dat', '10054.dat',
                      '10055.dat', '10056.dat', '10057.dat', '10061.dat', '10062.dat', '10066.dat', '10067.dat',
                      '10070.dat', '10071.dat', '10072.dat', '10073.dat', '10074.dat', '10075.dat', '10076.dat',
                      '10077.dat', '10079.dat', '10082.dat', '10083.dat', '10084.dat', '10085.dat', '10087.dat',
                      '10088.dat', '10089.dat', '10090.dat', '10092.dat', '10093.dat', '10095.dat', '10096.dat',
                      '10097.dat', '10098.dat', '10099.dat', '10100.dat', '10101.dat', '10102.dat', '10104.dat',
                      '10105.dat', '10106.dat', '10107.dat', '10108.dat', '10109.dat', '10110.dat', '10111.dat',
                      '10112.dat', '10113.dat', '10114.dat', '10115.dat', '10116.dat', '10117.dat', '10118.dat',
                      '10119.dat', '10120.dat', '10121.dat', '10122.dat', '10123.dat', '10124.dat', '10125.dat',
                      '10126.dat', '10127.dat', '10128.dat', '10129.dat', '10130.dat', '10131.dat', '10132.dat',
                      '10133.dat', '10134.dat', '10135.dat', '10136.dat', '10137.dat', '10138.dat', '10139.dat',
                      '10140.dat', '10141.dat', '10142.dat', '10143.dat', '10144.dat', '10145.dat', '10146.dat',
                      '10147.dat', '10148.dat', '10149.dat', '10150.dat', '10151.dat', '10152.dat', '10153.dat',
                      '10154.dat', '10155.dat', '10156.dat', '10157.dat', '10158.dat', '10159.dat', '10160.dat',
                      '10161.dat', '10162.dat', '10163.dat', '10164.dat', '10165.dat', '10166.dat', '10167.dat',
                      '10168.dat', '10169.dat', '10170.dat', '10171.dat', '10172.dat', '10173.dat', '10174.dat',
                      '10175.dat', '10176.dat', '10177.dat', '10178.dat', '10179.dat', '10180.dat', '10181.dat',
                      '10182.dat', '10183.dat', '10184.dat', '10185.dat', '10186.dat', '10187.dat', '10188.dat',
                      '10189.dat', '10190.dat', '10191.dat', '10192.dat', '10193.dat', '10194.dat', '10195.dat',
                      '10196.dat', '10197.dat', '10198.dat', '10199.dat', '10200.dat', '10201.dat', '10202.dat',
                      '10203.dat', '10204.dat', '10205.dat', '10206.dat', '10207.dat', '10208.dat', '10209.dat',
                      '10210.dat', '10211.dat', '10212.dat', '10213.dat', '10214.dat', '10215.dat', '10216.dat',
                      '10217.dat', '10218.dat', '10219.dat', '10220.dat', '10221.dat', '10222.dat', '10223.dat',
                      '10224.dat', '10225.dat', '10226.dat', '10227.dat', '10228.dat', '10229.dat', '10230.dat',
                      '10231.dat', '10232.dat', '10233.dat', '10234.dat', '10235.dat', '10236.dat', '10237.dat',
                      '10238.dat', '10239.dat', '10240.dat', '10241.dat', '10242.dat', '10243.dat', '10244.dat',
                      '10245.dat', '10246.dat', '10247.dat', '10248.dat', '10249.dat', '10250.dat', '10251.dat',
                      '10252.dat', '10253.dat', '10254.dat', '10255.dat', '10256.dat', '10257.dat', '10258.dat',
                      '10259.dat', '10260.dat', '10261.dat', '10262.dat', '10263.dat', '10264.dat', '10265.dat',
                      '10266.dat', '10267.dat', '10268.dat', '10269.dat', '10270.dat', '10271.dat', '10272.dat',
                      '10273.dat', '10274.dat', '10275.dat', '10276.dat', '10277.dat', '10278.dat', '10279.dat',
                      '10280.dat', '10281.dat', '10282.dat', '10283.dat', '10284.dat', '10285.dat', '10286.dat',
                      '10287.dat', '10288.dat', '10289.dat', '10290.dat', '10291.dat', '10292.dat', '10293.dat',
                      '10294.dat', '10295.dat', '10296.dat', '10297.dat', '10298.dat', '10299.dat', '10300.dat',
                      '10301.dat', '10302.dat', '10303.dat', '10305.dat', '10306.dat', '10307.dat', '10308.dat',
                      '10309.dat', '10311.dat', '10312.dat', '10313.dat', '10314.dat', '10315.dat', '10316.dat',
                      '10317.dat', '10318.dat', '10319.dat', '10320.dat', '10321.dat', '10322.dat', '10323.dat',
                      '10324.dat', '10325.dat', '10326.dat', '10327.dat', '10328.dat', '10329.dat', '10330.dat',
                      '10331.dat', '10332.dat', '10333.dat', '10334.dat', '10335.dat', '10336.dat', '10337.dat',
                      '10338.dat', '10339.dat', '10340.dat', '10341.dat', '10342.dat', '10343.dat', '10344.dat',
                      '10345.dat', '10346.dat', '10347.dat', '10348.dat', '10349.dat', '10350.dat', '10351.dat',
                      '10352.dat', '10353.dat', '10354.dat', '10355.dat', '10356.dat', '10357.dat', '10358.dat',
                      '10359.dat', '10360.dat', '10361.dat', '10362.dat', '10363.dat', '10364.dat', '10365.dat',
                      '10366.dat', '10367.dat', '10368.dat', '10369.dat', '10370.dat', '10371.dat', '10372.dat',
                      '10373.dat', '10374.dat', '10375.dat', '10376.dat', '10377.dat', '10378.dat', '10379.dat',
                      '10380.dat', '10381.dat', '10382.dat', '10383.dat', '10384.dat', '11382.dat', '11383.dat',
                      '11384.dat', '11385.dat', '11386.dat', '11387.dat', '11388.dat', '11389.dat', '11390.dat',
                      '11391.dat', '11392.dat', '11393.dat', '11394.dat', '11395.dat', '11396.dat', '11397.dat',
                      '11398.dat', '11399.dat', '11400.dat', '11401.dat', '11402.dat', '11403.dat', '11404.dat',
                      '11405.dat', '11406.dat', '11407.dat', '11408.dat', '11409.dat', '11410.dat', '11411.dat',
                      '11412.dat', '11414.dat', '11415.dat', '11416.dat', '11417.dat', '11418.dat', '11419.dat',
                      '11420.dat', '11421.dat', '11422.dat', '11423.dat', '11424.dat', '11425.dat', '11426.dat',
                      '11427.dat', '11428.dat', '11429.dat', '11430.dat', '11431.dat', '11432.dat', '11433.dat',
                      '11434.dat', '11435.dat', '11436.dat', '11437.dat', '11438.dat', '11439.dat', '11440.dat',
                      '11441.dat', '11442.dat', '11443.dat', '11444.dat', '11445.dat', '11446.dat', '11448.dat',
                      '11449.dat', '11450.dat', '11451.dat', '11452.dat', '11453.dat', '11454.dat', '11455.dat',
                      '11456.dat', '11457.dat', '11458.dat', '11459.dat', '11460.dat', '11461.dat', '11462.dat',
                      '11463.dat', '11464.dat', '11465.dat', '11466.dat', '11467.dat', '11468.dat', '11469.dat',
                      '11470.dat', '11471.dat', '11472.dat', '11473.dat', '11474.dat', '11475.dat', '11476.dat',
                      '11477.dat', '11478.dat', '11479.dat', '11480.dat', '11481.dat', '11482.dat', '11483.dat',
                      '11484.dat', '11485.dat', '11486.dat', '11487.dat', '11488.dat', '11489.dat', '11490.dat',
                      '11491.dat', '11492.dat', '11493.dat', '11494.dat', '11495.dat', '11496.dat', '11497.dat',
                      '11498.dat', '11499.dat', '11500.dat', '11501.dat', '11502.dat', '11503.dat', '11504.dat',
                      '11505.dat', '11506.dat', '11507.dat', '11508.dat', '11509.dat', '11510.dat', '11511.dat',
                      '11512.dat', '11513.dat', '11514.dat', '11515.dat', '11516.dat', '11517.dat', '11518.dat',
                      '11519.dat', '11520.dat', '11521.dat', '11522.dat', '11523.dat', '11524.dat', '11525.dat',
                      '11526.dat', '11527.dat', '11528.dat', '11529.dat', '11530.dat', '11531.dat', '11532.dat',
                      '11533.dat', '11534.dat', '11535.dat', '11536.dat', '11537.dat', '11538.dat', '11539.dat',
                      '11541.dat', '11542.dat', '11543.dat', '11544.dat', '11545.dat', '11546.dat', '11547.dat',
                      '11548.dat', '11549.dat', '11550.dat', '11551.dat', '11552.dat', '11553.dat', '11554.dat',
                      '11555.dat', '11556.dat', '11557.dat', '11558.dat', '11559.dat', '11560.dat', '11561.dat',
                      '11562.dat', '11563.dat', '11564.dat', '11565.dat', '11566.dat', '11567.dat', '11568.dat',
                      '11569.dat', '11570.dat', '11571.dat', '11572.dat', '11573.dat', '11574.dat', '11575.dat',
                      '11576.dat', '11577.dat', '11578.dat', '11579.dat', '11580.dat', '11581.dat', '11582.dat',
                      '11583.dat', '11584.dat', '11585.dat', '11586.dat', '11587.dat', '11588.dat', '11589.dat',
                      '11590.dat', '11591.dat', '11592.dat', '11593.dat', '11594.dat', '11595.dat', '11596.dat',
                      '11597.dat', '11598.dat', '11599.dat', '11600.dat', '11601.dat', '11602.dat', '11603.dat',
                      '11604.dat', '11605.dat', '11606.dat', '11607.dat', '11608.dat', '11609.dat', '11610.dat',
                      '11611.dat', '11612.dat', '11613.dat', '11614.dat', '11615.dat', '11616.dat', '11617.dat',
                      '11618.dat', '11619.dat', '11620.dat', '11621.dat', '11622.dat', '11623.dat', '11624.dat',
                      '11625.dat', '11626.dat', '11627.dat', '11628.dat', '11629.dat', '11630.dat', '11631.dat',
                      '11632.dat', '11633.dat', '11634.dat', '11635.dat', '11636.dat', '11637.dat', '11638.dat',
                      '11639.dat', '11640.dat', '11641.dat', '11642.dat', '11643.dat', '11644.dat', '11645.dat',
                      '11646.dat', '11647.dat', '11648.dat', '11649.dat', '11650.dat', '11651.dat', '11652.dat',
                      '11653.dat', '11654.dat', '11655.dat', '11656.dat', '11657.dat', '11658.dat', '11659.dat',
                      '11660.dat', '11661.dat', '11662.dat', '11663.dat', '11664.dat', '11665.dat', '11666.dat',
                      '11667.dat', '11668.dat', '11669.dat', '11670.dat', '11671.dat', '11672.dat', '11673.dat',
                      '11674.dat', '11675.dat', '11676.dat', '11677.dat', '11678.dat', '11679.dat', '11680.dat',
                      '11681.dat', '11682.dat', '11683.dat', '11684.dat', '11685.dat', '11686.dat', '11687.dat',
                      '11688.dat', '11689.dat', '11690.dat', '11691.dat', '11692.dat', '11693.dat', '11694.dat',
                      '11695.dat', '11696.dat', '11697.dat', '11698.dat', '11699.dat', '11700.dat', '11701.dat',
                      '11702.dat', '11703.dat', '11704.dat', '11705.dat', '11706.dat']

    lecteur = DatFileReader()
    for invoice in liste_invoices:
        contenu = lecteur.lire_texte_simple(f'invoice/{invoice}')

        # Traiter la facture
        parser = FactureParser(contenu[0])
        resume = {
            'informations_client': parser.obtenir_infos_client(),
            'articles': parser.obtenir_articles(),
            'totaux': parser.obtenir_totaux(),
            'notes': parser.obtenir_notes()
        }
        pairs = dict(pair.split('=') for pair in contenu[0].split('&'))
        client = pairs.get("Customer", "")
        date = pairs.get("Date", "")
    # Créer un nouveau classeur Excel
    writer = pd.ExcelWriter(nom_fichier, engine='openpyxl')

    # Créer le DataFrame pour les informations client
    df_client = pd.DataFrame([resume['informations_client']])
    df_client.to_excel(writer, sheet_name=date, startrow=1, header=True, index=False)

    # Créer le DataFrame pour les articles
    df_articles = pd.DataFrame(resume['articles'])
    df_articles.to_excel(writer, sheet_name=date, startrow=6, header=True, index=False)

    # Obtenir la feuille de calcul
    workbook = writer.book
    worksheet = writer.sheets[date]

    # Styles
    header_style = Font(bold=True, size=12)
    cell_border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # Formater l'en-tête
    worksheet['A1'] = 'FACTURE'
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet.merge_cells('A1:F1')
    worksheet['A1'].alignment = Alignment(horizontal='center')

    # Formater les informations client
    for col in range(1, df_client.shape[1] + 1):
            cell = worksheet.cell(row=2, column=col)
            cell.font = header_style
            cell.fill = header_fill
            cell.border = cell_border

    # Formater les articles
    for col in range(1, df_articles.shape[1] + 1):
            cell = worksheet.cell(row=7, column=col)
            cell.font = header_style
            cell.fill = header_fill
            cell.border = cell_border

    # Ajouter les totaux
    row_totaux = 8 + len(resume['articles'])
    worksheet.cell(row=row_totaux, column=1, value='Total HT:')
    worksheet.cell(row=row_totaux, column=2, value=resume['totaux']['total_ht'])
    worksheet.cell(row=row_totaux + 1, column=1, value='Montant payé:')
    worksheet.cell(row=row_totaux + 1, column=2, value=resume['totaux']['montant_paye'])

    # Ajouter les notes
    row_notes = row_totaux + 3
    worksheet.cell(row=row_notes, column=1, value='Notes:')
    worksheet.cell(row=row_notes, column=2, value=resume['notes']['notes_communes'])
    worksheet.cell(row=row_notes + 1, column=1, value='Remarques:')
    worksheet.cell(row=row_notes + 1, column=2, value=resume['notes']['notes_pied'])

    # Ajuster la largeur des colonnes
    for col in worksheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = max_length + 2

    # Sauvegarder le fichier
    writer._save()


# Example d'utilisation
# donnees_facture = "Customer=KAMEL%20ALG&CustomFieldsTopCount=0&CustomFieldsBottomCount=0&Date=2021-04-21&PaymentTerms=1&PaymentTermsDays=30&SalesPerson=&TaxRate=0&Address=&ShippingCosts=0&ShippingTaxName=None&ShippingTaxRate=0.000000%2F0.000000%2F0&ShippingAddress=&ShippingAddressRequired=0&ShipBy=&TrackingRef=&CustomerPO=&UseOldTaxRates=0&DefaultTaxName=None&DefaultCombinedTaxName=&DefaultTaxRate=0.000000%2F0.000000%2F0&Item1Description=POLYAMIDE&Item1Qty=20.000000&Item1Code=T%2012&Item1Discount=0.000000&Item1TaxRate=0.000000%2F0.000000%2F0&Item1TaxName=None&Item1CombinedTaxName=None&Item1UnitValue=25000&Item1UnitValueDigits=2&Item1VATPercentage=0%25&Item1VATAmount=0&Item2Description=POLYAMIDE&Item2Qty=20.000000&Item2Code=T%2014&Item2Discount=0.000000&Item2TaxRate=0.000000%2F0.000000%2F0&Item2TaxName=None&Item2CombinedTaxName=None&Item2UnitValue=25000&Item2UnitValueDigits=2&Item2VATPercentage=0%25&Item2VATAmount=0&Item3Description=POLYAMIDE&Item3Qty=20.000000&Item3Code=T%2016&Item3Discount=0.000000&Item3TaxRate=0.000000%2F0.000000%2F0&Item3TaxName=None&Item3CombinedTaxName=None&Item3UnitValue=25000&Item3UnitValueDigits=2&Item3VATPercentage=0%25&Item3VATAmount=0&ItemCount=3&IsTaxInclusive=0&TaxAmountCount=0&TaxAmountCountCombined=0&Total=1500000&Notes=&NotesIternal=&CommonNotes=Veuillez%20nous%20contacter%20pour%20plus%20d%27informations%20sur%20les%20options%20de%20paiement.&CommonFootNotes=Nous%20vous%20remercions%20de%20votre%20confiance.&AmountPaid=1500000&IsDraft=0&AmountRefunded=0"

# Liste des fichiers trouvés

liste_invoices =['10000.dat', '10001.dat', '10002.dat', '10003.dat', '10004.dat', '10005.dat', '10006.dat', '10007.dat', '10008.dat', '10009.dat', '10010.dat', '10011.dat', '10012.dat', '10013.dat', '10014.dat', '10015.dat', '10016.dat', '10017.dat', '10018.dat', '10019.dat', '10020.dat', '10021.dat', '10022.dat', '10023.dat', '10024.dat', '10025.dat', '10027.dat', '10028.dat', '10030.dat', '10032.dat', '10036.dat', '10037.dat', '10038.dat', '10039.dat', '10040.dat', '10041.dat', '10042.dat', '10043.dat', '10044.dat', '10045.dat', '10046.dat', '10047.dat', '10048.dat', '10049.dat', '10050.dat', '10051.dat', '10052.dat', '10053.dat', '10054.dat', '10055.dat', '10056.dat', '10057.dat', '10061.dat', '10062.dat', '10066.dat', '10067.dat', '10070.dat', '10071.dat', '10072.dat', '10073.dat', '10074.dat', '10075.dat', '10076.dat', '10077.dat', '10079.dat', '10082.dat', '10083.dat', '10084.dat', '10085.dat', '10087.dat', '10088.dat', '10089.dat', '10090.dat', '10092.dat', '10093.dat', '10095.dat', '10096.dat', '10097.dat', '10098.dat', '10099.dat', '10100.dat', '10101.dat', '10102.dat', '10104.dat', '10105.dat', '10106.dat', '10107.dat', '10108.dat', '10109.dat', '10110.dat', '10111.dat', '10112.dat', '10113.dat', '10114.dat', '10115.dat', '10116.dat', '10117.dat', '10118.dat', '10119.dat', '10120.dat', '10121.dat', '10122.dat', '10123.dat', '10124.dat', '10125.dat', '10126.dat', '10127.dat', '10128.dat', '10129.dat', '10130.dat', '10131.dat', '10132.dat', '10133.dat', '10134.dat', '10135.dat', '10136.dat', '10137.dat', '10138.dat', '10139.dat', '10140.dat', '10141.dat', '10142.dat', '10143.dat', '10144.dat', '10145.dat', '10146.dat', '10147.dat', '10148.dat', '10149.dat', '10150.dat', '10151.dat', '10152.dat', '10153.dat', '10154.dat', '10155.dat', '10156.dat', '10157.dat', '10158.dat', '10159.dat', '10160.dat', '10161.dat', '10162.dat', '10163.dat', '10164.dat', '10165.dat', '10166.dat', '10167.dat', '10168.dat', '10169.dat', '10170.dat', '10171.dat', '10172.dat', '10173.dat', '10174.dat', '10175.dat', '10176.dat', '10177.dat', '10178.dat', '10179.dat', '10180.dat', '10181.dat', '10182.dat', '10183.dat', '10184.dat', '10185.dat', '10186.dat', '10187.dat', '10188.dat', '10189.dat', '10190.dat', '10191.dat', '10192.dat', '10193.dat', '10194.dat', '10195.dat', '10196.dat', '10197.dat', '10198.dat', '10199.dat', '10200.dat', '10201.dat', '10202.dat', '10203.dat', '10204.dat', '10205.dat', '10206.dat', '10207.dat', '10208.dat', '10209.dat', '10210.dat', '10211.dat', '10212.dat', '10213.dat', '10214.dat', '10215.dat', '10216.dat', '10217.dat', '10218.dat', '10219.dat', '10220.dat', '10221.dat', '10222.dat', '10223.dat', '10224.dat', '10225.dat', '10226.dat', '10227.dat', '10228.dat', '10229.dat', '10230.dat', '10231.dat', '10232.dat', '10233.dat', '10234.dat', '10235.dat', '10236.dat', '10237.dat', '10238.dat', '10239.dat', '10240.dat', '10241.dat', '10242.dat', '10243.dat', '10244.dat', '10245.dat', '10246.dat', '10247.dat', '10248.dat', '10249.dat', '10250.dat', '10251.dat', '10252.dat', '10253.dat', '10254.dat', '10255.dat', '10256.dat', '10257.dat', '10258.dat', '10259.dat', '10260.dat', '10261.dat', '10262.dat', '10263.dat', '10264.dat', '10265.dat', '10266.dat', '10267.dat', '10268.dat', '10269.dat', '10270.dat', '10271.dat', '10272.dat', '10273.dat', '10274.dat', '10275.dat', '10276.dat', '10277.dat', '10278.dat', '10279.dat', '10280.dat', '10281.dat', '10282.dat', '10283.dat', '10284.dat', '10285.dat', '10286.dat', '10287.dat', '10288.dat', '10289.dat', '10290.dat', '10291.dat', '10292.dat', '10293.dat', '10294.dat', '10295.dat', '10296.dat', '10297.dat', '10298.dat', '10299.dat', '10300.dat', '10301.dat', '10302.dat', '10303.dat', '10305.dat', '10306.dat', '10307.dat', '10308.dat', '10309.dat', '10311.dat', '10312.dat', '10313.dat', '10314.dat', '10315.dat', '10316.dat', '10317.dat', '10318.dat', '10319.dat', '10320.dat', '10321.dat', '10322.dat', '10323.dat', '10324.dat', '10325.dat', '10326.dat', '10327.dat', '10328.dat', '10329.dat', '10330.dat', '10331.dat', '10332.dat', '10333.dat', '10334.dat', '10335.dat', '10336.dat', '10337.dat', '10338.dat', '10339.dat', '10340.dat', '10341.dat', '10342.dat', '10343.dat', '10344.dat', '10345.dat', '10346.dat', '10347.dat', '10348.dat', '10349.dat', '10350.dat', '10351.dat', '10352.dat', '10353.dat', '10354.dat', '10355.dat', '10356.dat', '10357.dat', '10358.dat', '10359.dat', '10360.dat', '10361.dat', '10362.dat', '10363.dat', '10364.dat', '10365.dat', '10366.dat', '10367.dat', '10368.dat', '10369.dat', '10370.dat', '10371.dat', '10372.dat', '10373.dat', '10374.dat', '10375.dat', '10376.dat', '10377.dat', '10378.dat', '10379.dat', '10380.dat', '10381.dat', '10382.dat', '10383.dat', '10384.dat', '11382.dat', '11383.dat', '11384.dat', '11385.dat', '11386.dat', '11387.dat', '11388.dat', '11389.dat', '11390.dat', '11391.dat', '11392.dat', '11393.dat', '11394.dat', '11395.dat', '11396.dat', '11397.dat', '11398.dat', '11399.dat', '11400.dat', '11401.dat', '11402.dat', '11403.dat', '11404.dat', '11405.dat', '11406.dat', '11407.dat', '11408.dat', '11409.dat', '11410.dat', '11411.dat', '11412.dat', '11414.dat', '11415.dat', '11416.dat', '11417.dat', '11418.dat', '11419.dat', '11420.dat', '11421.dat', '11422.dat', '11423.dat', '11424.dat', '11425.dat', '11426.dat', '11427.dat', '11428.dat', '11429.dat', '11430.dat', '11431.dat', '11432.dat', '11433.dat', '11434.dat', '11435.dat', '11436.dat', '11437.dat', '11438.dat', '11439.dat', '11440.dat', '11441.dat', '11442.dat', '11443.dat', '11444.dat', '11445.dat', '11446.dat', '11448.dat', '11449.dat', '11450.dat', '11451.dat', '11452.dat', '11453.dat', '11454.dat', '11455.dat', '11456.dat', '11457.dat', '11458.dat', '11459.dat', '11460.dat', '11461.dat', '11462.dat', '11463.dat', '11464.dat', '11465.dat', '11466.dat', '11467.dat', '11468.dat', '11469.dat', '11470.dat', '11471.dat', '11472.dat', '11473.dat', '11474.dat', '11475.dat', '11476.dat', '11477.dat', '11478.dat', '11479.dat', '11480.dat', '11481.dat', '11482.dat', '11483.dat', '11484.dat', '11485.dat', '11486.dat', '11487.dat', '11488.dat', '11489.dat', '11490.dat', '11491.dat', '11492.dat', '11493.dat', '11494.dat', '11495.dat', '11496.dat', '11497.dat', '11498.dat', '11499.dat', '11500.dat', '11501.dat', '11502.dat', '11503.dat', '11504.dat', '11505.dat', '11506.dat', '11507.dat', '11508.dat', '11509.dat', '11510.dat', '11511.dat', '11512.dat', '11513.dat', '11514.dat', '11515.dat', '11516.dat', '11517.dat', '11518.dat', '11519.dat', '11520.dat', '11521.dat', '11522.dat', '11523.dat', '11524.dat', '11525.dat', '11526.dat', '11527.dat', '11528.dat', '11529.dat', '11530.dat', '11531.dat', '11532.dat', '11533.dat', '11534.dat', '11535.dat', '11536.dat', '11537.dat', '11538.dat', '11539.dat', '11541.dat', '11542.dat', '11543.dat', '11544.dat', '11545.dat', '11546.dat', '11547.dat', '11548.dat', '11549.dat', '11550.dat', '11551.dat', '11552.dat', '11553.dat', '11554.dat', '11555.dat', '11556.dat', '11557.dat', '11558.dat', '11559.dat', '11560.dat', '11561.dat', '11562.dat', '11563.dat', '11564.dat', '11565.dat', '11566.dat', '11567.dat', '11568.dat', '11569.dat', '11570.dat', '11571.dat', '11572.dat', '11573.dat', '11574.dat', '11575.dat', '11576.dat', '11577.dat', '11578.dat', '11579.dat', '11580.dat', '11581.dat', '11582.dat', '11583.dat', '11584.dat', '11585.dat', '11586.dat', '11587.dat', '11588.dat', '11589.dat', '11590.dat', '11591.dat', '11592.dat', '11593.dat', '11594.dat', '11595.dat', '11596.dat', '11597.dat', '11598.dat', '11599.dat', '11600.dat', '11601.dat', '11602.dat', '11603.dat', '11604.dat', '11605.dat', '11606.dat', '11607.dat', '11608.dat', '11609.dat', '11610.dat', '11611.dat', '11612.dat', '11613.dat', '11614.dat', '11615.dat', '11616.dat', '11617.dat', '11618.dat', '11619.dat', '11620.dat', '11621.dat', '11622.dat', '11623.dat', '11624.dat', '11625.dat', '11626.dat', '11627.dat', '11628.dat', '11629.dat', '11630.dat', '11631.dat', '11632.dat', '11633.dat', '11634.dat', '11635.dat', '11636.dat', '11637.dat', '11638.dat', '11639.dat', '11640.dat', '11641.dat', '11642.dat', '11643.dat', '11644.dat', '11645.dat', '11646.dat', '11647.dat', '11648.dat', '11649.dat', '11650.dat', '11651.dat', '11652.dat', '11653.dat', '11654.dat', '11655.dat', '11656.dat', '11657.dat', '11658.dat', '11659.dat', '11660.dat', '11661.dat', '11662.dat', '11663.dat', '11664.dat', '11665.dat', '11666.dat', '11667.dat', '11668.dat', '11669.dat', '11670.dat', '11671.dat', '11672.dat', '11673.dat', '11674.dat', '11675.dat', '11676.dat', '11677.dat', '11678.dat', '11679.dat', '11680.dat', '11681.dat', '11682.dat', '11683.dat', '11684.dat', '11685.dat', '11686.dat', '11687.dat', '11688.dat', '11689.dat', '11690.dat', '11691.dat', '11692.dat', '11693.dat', '11694.dat', '11695.dat', '11696.dat', '11697.dat', '11698.dat', '11699.dat', '11700.dat', '11701.dat', '11702.dat', '11703.dat', '11704.dat', '11705.dat', '11706.dat']


lecteur = DatFileReader()
for invoice in liste_invoices:
      contenu = lecteur.lire_texte_simple(f'invoice/{invoice}')

      # Traiter la facture
      parser = FactureParser(contenu[0])
      resume = {
            'informations_client': parser.obtenir_infos_client(),
            'articles': parser.obtenir_articles(),
            'totaux': parser.obtenir_totaux(),
            'notes': parser.obtenir_notes()
        }
      pairs = dict(pair.split('=') for pair in contenu[0].split('&'))
      client = pairs.get("Customer", "")


      # Sauvegarder en Excel
      sauvegarder_excel(resume, f"{client}-{invoice}.xlsx")