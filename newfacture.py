from openpyxl import load_workbook
from pathlib import Path
from template_facture import generate_invoice  # إذا كان template في ملف آخر

SOURCE_DIR = Path("invoices_downloaded/ABD_ALMALEK_BORJMNAYEL")
OUTPUT_DIR = Path("factures_converties")
OUTPUT_DIR.mkdir(exist_ok=True)

# -------------------------------
def extract_invoice_data(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # ⚠️ عدّل هذه الخلايا إذا لزم
    client = ws["A4"].value
    date = ws["B4"].value
    delai = ws["C4"].value or 30

    articles = []
    row = 7  # بداية جدول المنتجات

    while ws.cell(row=row, column=1).value:
        desc = ws.cell(row=row, column=1).value
        code = ws.cell(row=row, column=2).value
        qte = ws.cell(row=row, column=3).value
        prix = ws.cell(row=row, column=4).value
        remise = ws.cell(row=row, column=5).value or 0
        tva = ws.cell(row=row, column=6).value or 0

        articles.append([desc, code, qte, prix, remise, tva])
        row += 1

    return client, date, delai, articles

# -------------------------------
def main():
    for file in SOURCE_DIR.glob("*.xlsx"):
        print(f"🔄 Traitement : {file.name}")

        client, date, delai, articles = extract_invoice_data(file)

        output_file = OUTPUT_DIR / f"Facture_{client}_{date}.xlsx"

        generate_invoice(
            filename=output_file,
            client=client,
            invoice_date=str(date),
            delai_paiement=delai,
            articles=articles
        )

        print(f"✅ Générée : {output_file.name}")

if __name__ == "__main__":
    main()
