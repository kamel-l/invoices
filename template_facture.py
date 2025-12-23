from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# =================================================
# TEMPLATE DE FACTURE
# =================================================
def generate_invoice(
    filename,
    client,
    invoice_date,
    delai_paiement,
    articles,
    montant_paye=0,
    notes="Veuillez nous contacter pour plus d'informations sur les options de paiement.",
    remarques="Nous vous remercions de votre confiance."
):
    wb = Workbook()
    ws = wb.active
    ws.title = invoice_date

    # ---------------- STYLES ----------------
    green_title = PatternFill("solid", fgColor="2E7D32")
    blue_block = PatternFill("solid", fgColor="1F4E79")
    white_bold = Font(color="FFFFFF", bold=True)
    white = Font(color="FFFFFF")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # ---------------- TITLE ----------------
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Facture {invoice_date}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws["A1"].fill = green_title

    # ---------------- CLIENT INFO ----------------
    headers = ["client", "date", "delai_paiement", "adresse", "ref_commande"]
    values = [client, invoice_date, delai_paiement, "", ""]

    ws.append(headers)
    ws.append(values)

    for col in range(1, 6):
        for row in (3, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = blue_block
            cell.font = white_bold if row == 3 else white
            cell.border = border

    # ---------------- TABLE HEADER ----------------
    ws.append([])
    ws.append(["description", "code", "quantite", "prix_unitaire", "remise", "tva", "total"])

    for col in range(1, 8):
        cell = ws.cell(row=6, column=col)
        cell.fill = blue_block
        cell.font = white_bold
        cell.alignment = center
        cell.border = border

    # ---------------- ARTICLES ----------------
    row = 7
    total_ht = 0

    for art in articles:
        desc, code, qte, prix, remise, tva = art
        total = 2000
        total_ht += total

        values = [desc, code, qte, prix, remise, tva, total]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = blue_block
            cell.font = white
            cell.border = border
            cell.alignment = right if isinstance(val, (int, float)) else left

        row += 1

    # ---------------- TOTALS ----------------
    ws.cell(row=row, column=1, value="Total HT:")
    ws.cell(row=row, column=4, value=total_ht)

    ws.cell(row=row+1, column=1, value="Montant payé:")
    ws.cell(row=row+1, column=4, value=montant_paye)

    for r in (row, row+1):
        for c in (1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_block
            cell.font = white
            cell.border = border

    # ---------------- NOTES ----------------
    ws.cell(row=row+3, column=1, value="Notes:")
    ws.cell(row=row+3, column=2, value=notes)

    ws.cell(row=row+4, column=1, value="Remarques:")
    ws.cell(row=row+4, column=2, value=remarques)

    for r in (row+3, row+4):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_block
            cell.font = white
            cell.border = border

    # ---------------- COLUMN WIDTH ----------------
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filename)
    print(f"✅ Facture générée : {filename}")
